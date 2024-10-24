import os
from collections.abc import Iterable, Sequence
from typing import Any, TextIO

import openpyxl
import pandas as pd
import xlrd
from django.db import models
from openpyxl.worksheet.worksheet import Worksheet
from xlrd.sheet import Cell, Sheet

from .battery_exceptions import EmptyFileError, UnsupportedFileTypeError
from .mappings import MACCOR_COLUMN_MAPPING
from .parsing_engines_base import ParsingEngineBase


class MaccorParsingEngine(ParsingEngineBase):
    name = "Maccor"
    description = "Maccor XLS/XLSX"
    valid: list[tuple[str, str]] = [
        ("application/vnd.ms-excel", ".xls"),
        ("application/CDFV2", ".xls"),
        ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
        ("application/zip", ".xlsx"),
        ("text/plain", ".csv"),
        ("text/plain", ".txt"),
    ]
    mandatory_columns: dict[str, dict[str, str | tuple[str, str]]] = {
        "Rec#": dict(symbol="Rec", unit=("Unitless", "1")),
        "Cyc#": dict(symbol="Cyl", unit=("Unitless", "1")),
        "Step": dict(symbol="Ns changes", unit=("Unitless", "1")),
        "TestTime": dict(symbol="t", unit=("Time", "s")),
        "StepTime": dict(symbol="ts", unit=("Time", "s")),
        "Amp-hr": dict(symbol="Q-Q_0", unit=("Charge", "A·h")),
        "Watt-hr": dict(symbol="E", unit=("Energy", "W·h")),
        "Amps": dict(symbol="I", unit=("Current", "A")),
        "Volts": dict(symbol="V", unit=("Voltage", "V")),
        "Temp 1": dict(symbol="T", unit=("Temperature", "C")),
    }

    @classmethod
    def factory(cls, file_obj: TextIO) -> ParsingEngineBase:
        """Factory method for creating a parsing engine.
        Different functions/methods are used depending on the file extension.

        Args:
            file_obj (TextIO): File to parse.
        """
        column_name_mapping = MACCOR_COLUMN_MAPPING
        ext = os.path.splitext(file_obj.name)[1]
        file_metadata: dict[str, Any] | list[Any] = {}
        if ext in [".csv", ".txt"]:
            skip_rows = get_header_size_csv(file_obj, set(cls.mandatory_columns.keys()))
            data = load_maccor_data_csv(file_obj, skip_rows)
            file_metadata = get_file_header_csv(file_obj, skip_rows)
            return cls(file_obj, skip_rows, data, file_metadata, column_name_mapping)
        elif ext not in [".xls", ".xlsx"]:
            raise UnsupportedFileTypeError(
                f"Unknown file extension for Maccor parsing engine '{ext}'."
            )
        elif ext == ".xls":
            sheet, datemode = factory_xls(file_obj)
            if sheet.ncols < 1 or sheet.nrows < 2:
                raise EmptyFileError()
        elif ext == ".xlsx":
            sheet, datemode = factory_xlsx(file_obj)
            if sheet.max_column < 1 or sheet.max_row < 2:
                raise EmptyFileError()
        skip_rows = get_header_size_excel(sheet, set(cls.mandatory_columns.keys()))
        data = load_maccor_data_excel(file_obj, skip_rows)
        file_metadata = get_file_header_excel(sheet, skip_rows, datemode)
        return cls(file_obj, skip_rows, data, file_metadata, column_name_mapping)


def factory_xls(file_obj: TextIO) -> tuple[Sheet | Worksheet, int | None]:
    """Factory method for retrieving information specific for Maccor XLS files.

    Args:
        file_obj (TextIO): File to load.

    Returns:
        A tuple with a sheet object and the datemode of the workbook.
    """
    file_obj.seek(0)
    book = xlrd.open_workbook(file_contents=file_obj.read(), on_demand=True)
    return book.sheet_by_index(0), book.datemode


def factory_xlsx(file_obj: TextIO) -> tuple[Sheet | Worksheet, int | None]:
    """Factory method for retrieving information specific for Maccor XLSX files.

    Args:
        file_obj (TextIO): File to load.

    Returns:
        A tuple with a sheet object and the datemode of the workbook.
    """
    file_obj.seek(0)
    book = openpyxl.load_workbook(file_obj, read_only=True)
    return book.active, None


def get_file_header_excel(
    sheet: Sheet | Worksheet, skip_rows: int, datemode: int | None
) -> dict[str, Any]:
    """Extracts the header from the Maccor file.

    The header is spread across columns, so these need to be scan for the key/value
    pairs to extract.

    Args:
        sheet (Union[Sheet, Worksheet]): The Excel sheet to scan.
        skip_rows (int): Number of rows containing the header.
        datemode (Optional[int]): The encoding of dates in the excel file.

    Returns:
        Dict[str, Any]: A dictionary with the header information.
    """
    header: dict[str, Any] = {}
    for i, row in enumerate(sheet):
        if i == skip_rows:
            break

        current = 0
        while current < len(row):
            key, value, current = get_metadata_value(current, row, datemode)
            if key == "":
                continue
            header[key] = value

    return header


def get_file_header_csv(file_obj: models.FileField, skip_rows: int) -> list[str]:
    """Extracts the header from the Maccor csv or txt file.

    Args:
        file_obj (TextIO): File to load the data from.
        skip_rows (int): Location of the header, assumed equal to the number of rows to
            skip.

    Returns:
        A list of rows for the header, without the termination characters and
        empty lines.
    """
    if skip_rows == 0:
        return []
    with file_obj.open("r") as f:
        header = list(filter(len, (f.readline().rstrip() for _ in range(skip_rows))))
        if type(header[0]) is bytes:
            header = [i.decode("iso-8859-1") for i in header]
    return header


def get_header_size_excel(sheet: Sheet | Worksheet, columns: set) -> int:
    """Reads the file and determines the size of the header.

    Args:
        sheet (Union[Sheet, Worksheet]): The Excel sheet to scan.
        columns (Set): Iterable with the elements to check that would identify
            this as NOT a metadata row.

    Returns:
        int: The size of the header.
    """
    for i, row in enumerate(sheet):
        if not is_metadata_row(row, columns, "excel"):
            return i
    return 0


def get_header_size_csv(file_obj: models.FileField, columns: set) -> int:
    """Reads the file and determines the size of the header.

    Args:
        file_obj (TextIO): File to load the data from.
        columns (Set): Iterable with the elements to check that would identify
            this as NOT a metadata row.

    Returns:
        int: The size of the header.
    """
    file_obj.seek(0)
    with file_obj.open("r") as f:
        lines = iter([i.decode("iso-8859-1") for i in f.readlines()])
        for i, line in enumerate(lines):
            if not is_metadata_row(line, columns, "csv"):
                return i

    return 0


def load_maccor_data_excel(file_obj: TextIO, skip_rows: int) -> pd.DataFrame:
    """Loads the data as a Pandas data frame.

    Args:
        file_obj (TextIO): File to load.
        skip_rows (int): Location of the header, assumed equal to the number of rows to
            skip.

    Returns:
        pd.DataFrame: A pandas dataframe with all the data.
    """
    file_obj.seek(0)
    data = pd.read_excel(file_obj, sheet_name=None, header=skip_rows, index_col=None)
    if isinstance(data, dict):
        data = pd.concat(data.values(), ignore_index=True)
    return data


def load_maccor_data_csv(file_obj: models.FileField, skip_rows: int) -> pd.DataFrame:
    """Loads the data as a Pandas data frame.

    Args:
        file_obj (TextIO): File to load.
        skip_rows (int): Location of the header, assumed equal to the number of rows to
            skip.

    Returns:
        pd.DataFrame: A pandas dataframe with all the data.
    """
    kwargs = dict(skiprows=skip_rows, encoding="iso-8859-1", encoding_errors="replace")
    file_obj.open("r")
    try:
        file_obj.seek(0)
        data = pd.read_csv(file_obj, delimiter=",", **kwargs)
    except pd.errors.ParserError:
        try:
            file_obj.seek(0)
            data = pd.read_csv(file_obj, delimiter="\t", **kwargs)
        except pd.errors.ParserError as err:
            raise UnsupportedFileTypeError(err)

    if len(data.columns) == 1:
        file_obj.seek(0)
        data = pd.read_csv(file_obj, delimiter="\t", **kwargs)

    if len(data.columns) == 1:
        raise UnsupportedFileTypeError()

    return data


def clean_value(value: str) -> str:
    """Cleans up the string and trims special characters.

    Args:
        value (str): The string to clean

    Returns:
        str: The string cleaned.
    """
    return value.replace("''", "'").strip().rstrip("\0").strip()


def is_metadata_row(row: Iterable, indicators: Iterable, filetype: str) -> bool:
    """Checks if a row is a metadata row.

    Args:
        row (Iterable): Iterable with the row data to check
        indicators (Iterable): Iterable with the elements to check that would identify
            this as NOT a metadata row.
        filetype (Str): Either "excel" or "csv".

    Returns:
        bool: True if it is identified as a metadata row
    """
    if filetype == "excel":
        row_values = [y.value if hasattr(y, "value") else y for y in row]
        return not any(x in indicators for x in row_values)
    elif filetype == "csv":
        return not any(x in row for x in indicators)
    else:
        raise ValueError(f"Unrecognized filetype: {filetype}")


def get_metadata_value(
    idx: int, row: Sequence[Cell], datemode: int | None
) -> tuple[str, str, int]:
    """A wrapper for metadata value parsing, handling special cases.

    Args:
        idx (int): Index of the cell with the next key to check
        row (Sequence[Cell]): Row of cells
        datemode (int): The datemode the cell containing dates are using.

    Raises:
        ValueError: If the index of the next cell is not bigger than the current
            one.

    Returns:
        Tuple[str, str, int]: A tuple with the processed key, value and the next
        index to check.
    """
    key = row[idx]
    if not key:
        key = ""
    elif isinstance(key.value, str):
        key = key.value.replace("''", "'").strip().rstrip(":")
    else:
        key = key.value

    if hasattr(key, "__iter__") and "Date" in key:
        value = str(
            xlrd.xldate.xldate_as_datetime(row[idx + 1].value, datemode)
            if datemode is not None
            else row[idx + 1].value
        )
        new_idx = idx + 2
    elif hasattr(key, "__iter__") and "Procedure" in key:
        value = clean_value(row[idx + 1].value) + ", " + clean_value(row[idx + 2].value)
        new_idx = idx + 3
    else:
        value = row[idx + 1].value if idx + 1 < len(row) else ""
        new_idx = idx + 2

    if new_idx <= idx:
        raise ValueError("Non-increasing index number when processing metadata cells.")
    return key, value, new_idx
